import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import os

# =================配置区域=================
file_path = '板块回测汇总结果_含总笔数2026-03-12-13-57-01_test_30f_rps__bug版第三根有可能卖出.xlsx'
sheet_name = '所有交易明细'           
column_name = '实际盈亏比例'                
initial_capital = 100000              
num_simulations = 100                 
output_sheet = '随机复利模拟结果'
# ==========================================

try:
    # 1. 模拟计算逻辑
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    returns = df[column_name].dropna().values / 100 
    num_trades = len(returns) 
    random_indices = np.random.randint(0, len(returns), size=(num_simulations, num_trades))
    sim_returns = returns[random_indices]
    curves = initial_capital * np.cumprod(1 + sim_returns, axis=1)
    curves = np.insert(curves, 0, initial_capital, axis=1)

    # 2. 统计指标计算
    final_values = curves[:, -1]
    running_max = np.maximum.accumulate(curves, axis=1)
    drawdowns = (curves - running_max) / running_max
    max_drawdowns = drawdowns.min(axis=1)

    stats = [
        ("平均最终净值", f"{final_values.mean():,.2f}"),
        ("最高最终净值", f"{final_values.max():,.2f}"),
        ("最低最终净值", f"{final_values.min():,.2f}"),
        ("资产翻倍率", f"{np.sum(final_values > 200000)/num_simulations:.2%}"),
        ("资产减半率", f"{np.sum(final_values < 50000)/num_simulations:.2%}"),
        ("平均最大回撤", f"{max_drawdowns.mean():.2%}"),
        ("最极端回撤", f"{max_drawdowns.min():.2%}"),
        ("回撤>30%占比", f"{np.sum(max_drawdowns < -0.3)/num_simulations:.2%}"),
        ("回撤>50%占比", f"{np.sum(max_drawdowns < -0.5)/num_simulations:.2%}")
    ]

    # 3. 生成大图 (保存到本地)
    plt.figure(figsize=(30, 18)) 
    plt.plot(pd.DataFrame(curves.T), color='gray', alpha=0.15, linewidth=1.5)
    plt.plot(curves.mean(axis=0), color='red', linewidth=6, label='均值期望路径')
    plt.yscale('log')
    plt.title(f"Monte Carlo Simulation: {num_trades} Trades", fontsize=40)
    plt.grid(True, which="both", ls="--", alpha=0.5)
    
    temp_img = "monte_carlo_final_plot.png"
    plt.savefig(temp_img, dpi=120) 
    plt.close()

    # 4. 写入原始路径数据 (这次从 D 列开始写，留出 A-C 列放统计)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        path_df = pd.DataFrame(curves.T)
        path_df.columns = [f'路径_{i+1}' for i in range(num_simulations)]
        path_df.to_excel(writer, sheet_name=output_sheet, startrow=0, startcol=3) # startcol=3 是 D 列

    # 5. 【原生填值】在 A1 写入统计摘要
    wb = load_workbook(file_path)
    ws = wb[output_sheet]
    
    # 设置表头文字
    ws['A1'] = "策略统计指标"
    ws['B1'] = "数值结果"
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    
    # 填入数据
    for i, (label, val) in enumerate(stats, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)

    # 调整前两列宽度
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

    # 6. 插入大图到最右侧 (放在 100 条路径数据的右边)
    # 路径数据在 D 到 DB 列左右，图片放在 DE 以后
    img = Image(temp_img)
    img.width = 2400
    img.height = 1440
    ws.add_image(img, 'DE2') 
    
    wb.save(file_path)

    if os.path.exists(temp_img):
        os.remove(temp_img)

    print(f"模拟计算完成！")
    print(f"统计指标：已存入 A1:B10")
    print(f"原始数据：已存入 D 列以后")
    print(f"可视化大图：已存入 DE2 单元格以后（向右滑动可见）")

except Exception as e:
    print(f"程序运行出错: {e}")