import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import os

# =================配置区域=================
file_path = '板块回测汇总结果_含总笔数2026-03-13-13-15-27.xlsx'
sheet_name = '所有交易明细'           
column_name = '实际盈亏比例'                
initial_capital = 100000              
num_simulations = 100                 
output_sheet = '随机复利模拟结果动态收益'

# 动态风险控制参数
DD_THRESHOLD = 0.15  # 回撤阈值：10%
RISK_REDUCTION = 0.5  # 风险缩减系数：回撤后只承担 0.5 倍风险
# ==========================================

try:
    # 1. 加载数据
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    # 原始收益率（基于1%风险金）
    returns = df[column_name].dropna().values / 100 
    num_trades = len(returns) 

    # 准备存储所有模拟路径的矩阵
    # curves 形状: (模拟次数, 交易笔数 + 1)
    curves = np.zeros((num_simulations, num_trades + 1))
    curves[:, 0] = initial_capital

    # 随机生成每条路径对应的交易索引
    random_indices = np.random.randint(0, len(returns), size=(num_simulations, num_trades))

    # 2. 执行带有动态风控的模拟计算
    for s in range(num_simulations):
        current_equity = initial_capital
        high_water_mark = initial_capital
        
        for t in range(num_trades):
            # 计算当前回撤
            drawdown = (high_water_mark - current_equity) / high_water_mark
            
            # 确定风险系数
            risk_multiplier = RISK_REDUCTION if drawdown > DD_THRESHOLD else 1.0
            
            # 获取当前随机抽取的原始收益率
            raw_ret = returns[random_indices[s, t]]
            
            # 计算实际收益率并更新净值
            actual_ret = raw_ret * risk_multiplier
            current_equity *= (1 + actual_ret)
            
            # 更新路径数据
            curves[s, t + 1] = current_equity
            
            # 更新历史最高点
            if current_equity > high_water_mark:
                high_water_mark = current_equity

    # 3. 统计指标计算 (保持原逻辑不变)
    final_values = curves[:, -1]
    
    # 计算每条路径的最大回撤
    running_max = np.maximum.accumulate(curves, axis=1)
    drawdowns = (curves - running_max) / running_max
    max_drawdowns = np.min(drawdowns, axis=1)

    # 汇总指标
    stats = [
        ("平均最终净值", np.mean(final_values)),
        ("最高最终净值", np.max(final_values)),
        ("最低最终净值", np.min(final_values)),
        ("资产翻倍率", np.mean(final_values >= initial_capital * 2)),
        ("资产减半率", np.mean(final_values <= initial_capital * 0.5)),
        ("平均最大回撤", np.mean(max_drawdowns)),
        ("最极端回撤", np.min(max_drawdowns)),
        (f"回撤>{DD_THRESHOLD*100:.0f}%占比", np.mean(max_drawdowns <= -DD_THRESHOLD)),
        ("回撤>30%占比", np.mean(max_drawdowns <= -0.3)),
        ("回撤>50%占比", np.mean(max_drawdowns <= -0.5)),
    ]

    # 4. 可视化绘制
    plt.figure(figsize=(15, 8))
    for i in range(num_simulations):
        plt.plot(curves[i], color='gray', alpha=0.2, linewidth=0.5)
    
    plt.plot(np.mean(curves, axis=0), color='blue', label='Mean Path', linewidth=2)
    plt.axhline(initial_capital, color='red', linestyle='--', alpha=0.5)
    plt.title(f"Monte Carlo Simulation (Dynamic Risk: DD > {DD_THRESHOLD*100}% -> Risk x{RISK_REDUCTION})", fontsize=20)
    plt.xlabel("Number of Trades", fontsize=15)
    plt.ylabel("Equity", fontsize=15)
    plt.yscale('log') # 建议复利使用对数坐标
    plt.grid(True, which="both", ls="--", alpha=0.5)
    
    temp_img = "monte_carlo_dynamic_plot.png"
    plt.savefig(temp_img, dpi=120) 
    plt.close()

    # 5. 写入 Excel (保持原逻辑)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        path_df = pd.DataFrame(curves.T)
        path_df.columns = [f'路径_{i+1}' for i in range(num_simulations)]
        path_df.to_excel(writer, sheet_name=output_sheet, startrow=0, startcol=3)

    wb = load_workbook(file_path)
    ws = wb[output_sheet]
    ws['A1'] = "策略统计指标 (动态风控模式)"
    ws['B1'] = "数值结果"
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    
    for i, (label, val) in enumerate(stats, start=2):
        ws.cell(row=i, column=1, value=label)
        if "率" in label or "占比" in label or "回撤" in label:
            ws.cell(row=i, column=2, value=val).number_format = '0.00%'
        else:
            ws.cell(row=i, column=2, value=val).number_format = '#,##0.00'

    img = Image(temp_img)
    ws.add_image(img, 'A15')
    wb.save(file_path)
    print(f"成功！动态风险管理模拟已完成，结果保存至 {file_path}")

except Exception as e:
    print(f"运行出错: {e}")