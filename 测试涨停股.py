import pandas as pd
from pytdx.hq import TdxHq_API
import redis
import time
import schedule
import os
from datetime import datetime
import gupiaojichu
import logging
import json
import winsound
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from collections import defaultdict
import logging


# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('stock_data_1.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

class StockDataCollector:
    def __init__(self, blk_file_path):
        self.blk_file_path = blk_file_path
        self.stock_list = self.load_stock_list()
        self.profit_records = {}  # 存储收益记录，格式: {日期: [{股票信息}, ...]}
        self.servers = [('36.153.42.16', 7709)]
        logging.info(f"初始化完成，共加载 {len(self.stock_list)} 只股票")
    
    def load_stock_list(self):
        stock_list = []
        try:
            with open(self.blk_file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            for line in lines[1:]:
                line = line.strip()
                if line:
                    market_code = line[0]
                    stock_code = line[1:7]
                    if market_code == '0':
                        market = 0
                        market_prefix = 'sz'
                    else:
                        market = 1
                        market_prefix = 'sh'
                    full_code = f"{market_prefix}{stock_code}"
                    stock_list.append((market, stock_code, full_code))
        except Exception as e:
            logging.error(f"读取blk文件失败: {e}")
        return stock_list
    
    def get_5min_data(self, market, stock_code, full_code):
        api = TdxHq_API()
        for server_ip, server_port in self.servers:
            try:
                if api.connect(server_ip, server_port):
                    data = api.get_security_bars(9, market, stock_code, 0, 500)
                    api.disconnect()
                    if data:
                        result_list = []
                        result_list_high = []
                        result_list_low = []
                        for bar in data:
                            result = {
                                'open': float(bar['open']),
                                'high': float(bar['high']),
                                'low': float(bar['low']),
                                'close': float(bar['close']),
                                'volume': float(bar['vol']),
                                'datetime': bar['datetime']  # 假设格式为'YYYYMMDDHHMM'
                            }
                            result_list_high.append(float(bar['high']))
                            result_list_low.append(float(bar['low']))
                            result_list.append(result)
                        return result_list, result_list_high, result_list_low
                    else:
                        logging.warning(f"未获取到 {full_code} 的数据")
                        return None, None, None
                else:
                    logging.debug(f"连接服务器 {server_ip}:{server_port} 失败")
            except Exception as e:
                logging.debug(f"服务器获取 {full_code} 出错: {e}")
                continue
        logging.error(f"所有服务器都无法获取 {full_code} 的数据")
        return None, None, None
    


    def update_all_stocks(self):
        self.stock_list = self.load_stock_list()
        logging.info("开始更新所有股票数据...")
        
        for market, stock_code, full_code in self.stock_list:
            try:
                stock_data, _, _ = self.get_5min_data(market, stock_code, full_code)
                if stock_data and len(stock_data) >= 11:  # 确保有足够K线（需到次K的次K）
                    # 计算ZT和YZB状态
                    for i in range(1, len(stock_data)):
                        current = stock_data[i]
                        prev = stock_data[i-1]
                        price_ratio = current['close'] / prev['close']
                        current['ZT'] = (price_ratio >= 1.095) and (current['low'] < current['high'])
                        current['YZB'] = (price_ratio >= 1.095) and \
                                        (current['open'] == current['high'] == current['close'] == current['low'])
                        
                    for i in range(1, len(stock_data)):
                        latest_idx = i
                        latest_k = stock_data[latest_idx]
                        if not latest_k.get('ZT', False):
                            continue

                        # 条件判断函数（保持不变）
                        def check_cond1():
                            if latest_idx < 1: return False
                            prev1 = stock_data[latest_idx - 1]
                            return prev1.get('ZT') and (latest_k['volume'] < prev1['volume'])
                        def check_cond2():
                            if latest_idx < 2: return False
                            prev1, prev2 = stock_data[latest_idx-1], stock_data[latest_idx-2]
                            return prev1.get('YZB') and prev2.get('ZT') and (latest_k['volume'] < prev2['volume'])
                        def check_cond3():
                            if latest_idx < 3: return False
                            prev1, prev2, prev3 = [stock_data[latest_idx-i] for i in range(1,4)]
                            return all([prev1.get('YZB'), prev2.get('YZB')]) and prev3.get('ZT') and (latest_k['volume'] < prev3['volume'])
                        def check_cond4():
                            if latest_idx < 4: return False
                            prevs_yzb = [stock_data[latest_idx - i] for i in range(1,4)]
                            prev4 = stock_data[latest_idx - 4]
                            return all(yzb.get('YZB') for yzb in prevs_yzb) and prev4.get('ZT') and (latest_k['volume'] < prev4['volume'])
                        def check_cond5():
                            if latest_idx < 5: return False
                            prevs_yzb = [stock_data[latest_idx - i] for i in range(1,5)]
                            prev5 = stock_data[latest_idx - 5]
                            return all(yzb.get('YZB') for yzb in prevs_yzb) and prev5.get('ZT') and (latest_k['volume'] < prev5['volume'])
                        def check_cond6():
                            if latest_idx < 6: return False
                            prevs_yzb = [stock_data[latest_idx - i] for i in range(1,6)]
                            prev6 = stock_data[latest_idx - 6]
                            return all(yzb.get('YZB') for yzb in prevs_yzb) and prev6.get('ZT') and (latest_k['volume'] < prev6['volume'])
                        def check_cond7():
                            if latest_idx < 7: return False
                            prevs_yzb = [stock_data[latest_idx - i] for i in range(1,7)]
                            prev7 = stock_data[latest_idx - 7]
                            return all(yzb.get('YZB') for yzb in prevs_yzb) and prev7.get('ZT') and (latest_k['volume'] < prev7['volume'])
                        def check_cond8():
                            if latest_idx < 8: return False
                            prevs_yzb = [stock_data[latest_idx - i] for i in range(1,8)]
                            prev8 = stock_data[latest_idx - 8]
                            return all(yzb.get('YZB') for yzb in prevs_yzb) and prev8.get('ZT') and (latest_k['volume'] < prev8['volume'])
                        def check_cond9():
                            if latest_idx < 9: return False
                            prevs_yzb = [stock_data[latest_idx - i] for i in range(1,9)]
                            prev9 = stock_data[latest_idx - 9]
                            return all(yzb.get('YZB') for yzb in prevs_yzb) and prev9.get('ZT') and (latest_k['volume'] < prev9['volume'])

                        # 满足任一条件时处理
                        if any([check_cond1(), check_cond2(), check_cond3(), check_cond4(),
                                check_cond5(), check_cond6(), check_cond7(), check_cond8(), check_cond9()]):
                            logging.info(f"{full_code} 满足选股条件！")

                            # 1. 定义关键K线索引
                            current_idx = latest_idx  # 今K
                            next_idx = current_idx + 1  # 次K
                            next_next_idx = current_idx + 2  # 次K的次K

                            # 检查K线是否存在
                            if next_next_idx >= len(stock_data):
                                logging.debug(f"{full_code} K线不足（需次K和次K的次K），跳过")
                                continue

                            # 2. 提取关键价格
                            current_k = stock_data[current_idx]
                            next_k = stock_data[next_idx]
                            next_next_k = stock_data[next_next_idx]
                            current_close = current_k['close']  # 今K涨停收盘价
                            next_open = next_k['open']  # 次K开盘价

                            # 3. 次K开盘价低于今K收盘价则略过
                            if next_open < current_close:
                                logging.debug(f"{full_code} 次K开盘({next_open}) < 今K收盘({current_close})，略过")
                                continue
                            if next_k['high'] < current_close+0.01:  # 涨幅>次K开盘价
                                    continue
                            # 4. 计算涨幅和跌幅
                            next_open_ratio = (next_open / current_close) - 1  # 次K开盘涨幅（相对今K）

                            # 5. 计算买入价
                            buy_price = current_close * 1.05  # 5%位置买入
                            if next_open_ratio >= 0.05:  # 次K开盘≥5%
                                if buy_price < next_k['low']:  # 跌幅<5%
                                    continue
                            else:  # 次K开盘<5%
                                buy_price = next_open  # 开盘价买入

                            # 6. 计算卖出价（次K的次K开盘价）
                            sell_price = next_next_k['open']

                            # 7. 计算收益（百分比）
                            profit_ratio = (sell_price - buy_price) / buy_price * 100

                            # 8. 提取日期（从当前K线datetime取前8位，如20231009）
                            trade_date = str(current_k['datetime'])[:10]

                            # 9. 存入记录字典
                            if trade_date not in self.profit_records:
                                self.profit_records[trade_date] = []
                            self.profit_records[trade_date].append({
                                '股票代码': full_code,
                                '收益': round(profit_ratio, 2)
                            })
                            logging.info(f"{full_code} 计算完成：收益={round(profit_ratio, 2)}%")

                else:
                    logging.warning(f"{full_code} 数据不足或获取失败")
            except Exception as e:
                logging.error(f"处理 {full_code} 时出错: {e}")

        # 10. 所有股票处理完后写入Excel
        if self.profit_records:
            try:
                # 转换为DataFrame
                all_records = []
                for date, stocks in self.profit_records.items():
                    for stock in stocks:
                        stock['日期'] = date
                        all_records.append(stock)
                date_data = defaultdict(lambda: {'stocks': [], 'profits': []})
                for record in all_records:
                    date = record['日期']
                    stock_name = record['股票代码']  # 确保键与实际数据一致
                    profit = record['收益']      # 确保键与实际数据一致
                    date_data[date]['stocks'].append(stock_name)
                    date_data[date]['profits'].append(profit)

                # 2. 按日期排序（保证列顺序）
                sorted_dates = sorted(date_data.keys())
                if not sorted_dates:
                    logging.info("无收益记录可写入")
                    exit()

                # 3. 确定最大行数（各日期中股票数量的最大值）
                max_rows = max(len(data['stocks']) for data in date_data.values())

                # 4. 使用openpyxl写入Excel并设置格式
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "收益记录"

                # 写入表头（第一行）
                current_col = 1  # 从第1列开始（A列），如需从E列开始可改为5
                for date in sorted_dates:
                    # 股票名列标题（日期）
                    ws.cell(row=1, column=current_col).value = date
                    ws.cell(row=1, column=current_col).alignment = Alignment(horizontal='center')  # 居中
                    
                    # 盈亏列标题（介入盈亏）
                    ws.cell(row=1, column=current_col + 1).value = "介入盈亏"
                    ws.cell(row=1, column=current_col + 1).alignment = Alignment(horizontal='center')  # 居中
                    
                    current_col += 2  # 移动到下一个日期的列

                # 写入数据行（从第二行开始）
                for row_idx in range(2, max_rows + 2):  # 行号从2开始，共max_rows行
                    data_idx = row_idx - 2  # 数据索引（0-based）
                    current_col = 1
                    
                    for date in sorted_dates:
                        stocks = date_data[date]['stocks']
                        profits = date_data[date]['profits']
                        
                        # 写入股票名称
                        if data_idx < len(stocks):
                            ws.cell(row=row_idx, column=current_col).value = stocks[data_idx]
                        else:
                            ws.cell(row=row_idx, column=current_col).value = ""  # 空值填充
                        
                        # 写入介入盈亏
                        if data_idx < len(profits):
                            ws.cell(row=row_idx, column=current_col + 1).value = profits[data_idx]
                        else:
                            ws.cell(row=row_idx, column=current_col + 1).value = ""  # 空值填充
                        
                        current_col += 2

                # 调整列宽（根据内容自适应）
                for col in range(1, ws.max_column + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12  # 固定列宽

                # 保存文件
                wb.save('涨停股收益记录.xlsx')
                logging.info("收益记录已写入 涨停股收益记录.xlsx")
            except Exception as e:
                logging.error(f"写入Excel失败: {e}")
        return


    def get_stock_data_from_redis(self, full_code):
        try:
            redis_key = f"shishi:{full_code}"
            data_list = self.redis_client.lrange(redis_key, 0, -1)
            result = []
            for data_json in data_list:
                data_item = json.loads(data_json)
                result.append(data_item)
            return result
        except Exception as e:
            logging.error(f"从Redis获取 {full_code} 数据失败: {e}")
            return []
    
    def run(self, interval_seconds=2):
        logging.info(f"开始定时数据收集，间隔: {interval_seconds}秒")
        self.update_all_stocks()
        # 如需定时运行可取消下面注释
        # schedule.every(interval_seconds).seconds.do(self.update_all_stocks)
        # try:
        #     while True:
        #         schedule.run_pending()
        #         time.sleep(1)
        # except KeyboardInterrupt:
        #     logging.info("程序被用户中断")


def main():
    BLOB_FILE_PATH = r"D:\zd_hbzq\T0002\blocknew\ZB.blk"
    UPDATE_INTERVAL = 1
    if not os.path.exists(BLOB_FILE_PATH):
        logging.error(f"blk文件不存在: {BLOB_FILE_PATH}")
        return
    collector = StockDataCollector(blk_file_path=BLOB_FILE_PATH)
    collector.run(interval_seconds=UPDATE_INTERVAL)


if __name__ == "__main__":
    main()



















# import pandas as pd
# from pytdx.hq import TdxHq_API
# import redis
# import time
# import schedule
# import os
# from datetime import datetime
# import gupiaojichu
# import logging
# import json
# import winsound
# import numpy as np

# # 配置日志
# logging.basicConfig(
#     level=logging.INFO,
#     format='%(asctime)s - %(levelname)s - %(message)s',
#     handlers=[
#         logging.FileHandler('stock_data_1.log', encoding='utf-8'),
#         logging.StreamHandler()
#     ]
# )

# class StockDataCollector:
#     def __init__(self, blk_file_path):
#         """
#         初始化股票数据收集器
        
#         Args:
#             blk_file_path: blk文件路径
#             redis_host: Redis主机
#             redis_port: Redis端口
#             redis_db: Redis数据库编号
#         """
#         self.blk_file_path = blk_file_path
#         self.stock_list = self.load_stock_list()
#         self.triggered_stocks = set()  # 新增：用于记录已触发的股票代码
#         # 服务器列表
#         self.servers = [
#             # ('152.136.167.10', 7709),
#             ('36.153.42.16', 7709)
#         ]
        
#         logging.info(f"初始化完成，共加载 {len(self.stock_list)} 只股票")
    
#     def load_stock_list(self):
#         """
#         从blk文件加载股票列表
        
#         Returns:
#             list: 股票代码列表，格式为 [('市场代码', '股票代码', '完整代码'), ...]
#         """
#         stock_list = []
        
#         try:
#             with open(self.blk_file_path, 'r', encoding='utf-8') as f:
#                 lines = f.readlines()
                
#             # 跳过第一行空白，从第二行开始处理
#             for line in lines[1:]:
#                 line = line.strip()
#                 if line:
#                     # 第一位是市场代码，后六位是股票代码
#                     market_code = line[0]
#                     stock_code = line[1:7]
                    
#                     # 将市场代码转换为pytdx需要的格式
#                     # 0: 深圳, 1: 上海
#                     if market_code == '0':
#                         market = 0  # 深圳
#                         market_prefix = 'sz'
#                     else:
#                         market = 1  # 上海
#                         market_prefix = 'sh'
                    
#                     full_code = f"{market_prefix}{stock_code}"
#                     stock_list.append((market, stock_code, full_code))
                    
#                     logging.debug(f"加载股票: {full_code}")
                    
#         except Exception as e:
#             logging.error(f"读取blk文件失败: {e}")
            
#         return stock_list
    
#     def get_5min_data(self, market, stock_code, full_code):
#         """
#         获取单只股票的5分钟K线数据（200条）
        
#         Args:
#             market: 市场代码 (0: 深圳, 1: 上海)
#             stock_code: 股票代码
#             full_code: 完整股票代码 (如 sh600000)
            
#         Returns:
#             list or None: 200条K线数据，每条包含open, close, high, low, datetime
#         """
#         api = TdxHq_API()
        
#         # 尝试所有服务器直到成功
#         for server_ip, server_port in self.servers:
#             try:
#                 if api.connect(server_ip, server_port):
#                     logging.debug(f"成功连接到服务器 {server_ip}:{server_port}，获取 {full_code}")
                    
#                     # 获取5分钟K线数据 (category=0)，获取200条
#                     data = api.get_security_bars(9, market, stock_code, 0, 500)
#                     api.disconnect()
                    
#                     if data:
#                         # 处理所有200条数据
#                         result_list = []
#                         result_list_high = []
#                         result_list_low = []
#                         for bar in data:
#                             # 提取需要的字段
#                             result = {
#                                 'open': float(bar['open']),
#                                 'high': float(bar['high']),
#                                 'low': float(bar['low']),
#                                 'close': float(bar['close']),
#                                 'volume': float(bar['vol']),  # 新增：成交量
#                                 'datetime': bar['datetime']
#                             }
#                             result_list_high.append(float(bar['high']))
#                             result_list_low.append(float(bar['low']))
#                             result_list.append(result)
                        
#                         # 按时间排序（从旧到新）
#                         # result_list.sort(key=lambda x: x['datetime'])
                        
#                         logging.debug(f"成功获取 {full_code} 数据: {len(result_list)} 条")
#                         return result_list,result_list_high,result_list_low
#                     else:
#                         logging.warning(f"未获取到 {full_code} 的数据")
#                         return None,None,None
#                 else:
#                     logging.debug(f"连接服务器 {server_ip}:{server_port} 失败")
                    
#             except Exception as e:
#                 logging.debug(f"服务器 {server_ip}:{server_port} 获取 {full_code} 出错: {e}")
#                 continue
                
#         logging.error(f"所有服务器都无法获取 {full_code} 的数据")
#         return None,None,None
    

                

#     def update_all_stocks(self):
#         """
#         更新所有股票数据到Redis
#         """
#         self.stock_list = self.load_stock_list()
        
#         logging.info("开始更新所有股票数据...")
        
#         for market, stock_code, full_code in self.stock_list:
#             try:
#                 # 获取股票数据（200条）
#                 stock_data,stock_data_high,stock_data_low = self.get_5min_data(market, stock_code, full_code)
                
#                 if stock_data:
#                                         # 确保有足够的K线数量（COND9需要用到前9根数据）
#                     if len(stock_data) < 10:
#                         logging.debug(f"{full_code} K线数量不足（{len(stock_data)}条），无法判断条件")
#                         continue

#                     # 计算每根K线的ZT（普通涨停）和YZB（一字板）状态
#                     for i in range(1, len(stock_data)):  # 从第2根K线开始（需要前一根数据）
#                         current = stock_data[i]
#                         prev = stock_data[i-1]
#                         # 涨幅计算：当前收盘价 / 前收盘价
#                         price_ratio = current['close'] / prev['close']
                        
#                         # 普通涨停：涨幅≥9.5% 且 最低价<最高价（有成交波动）
#                         current['ZT'] = (price_ratio >= 1.095) and (current['low'] < current['high'])
                        
#                         # 一字板：涨幅≥9.5% 且 开盘价=最高价=收盘价=最低价（无波动）
#                         current['YZB'] = (price_ratio >= 1.095) and \
#                                         (current['open'] == current['high']) and \
#                                         (current['open'] == current['close']) and \
#                                         (current['low'] == current['high'])

#                     # 检查最新一根K线是否满足条件
#                     latest_idx = len(stock_data) - 1
#                     latest_k = stock_data[latest_idx]
                    
#                     # 当前K线必须是普通涨停才继续判断
#                     if not latest_k.get('ZT', False):
#                         continue

#                     # 定义各条件判断函数
#                     def check_cond1():
#                         # COND1: 当前ZT + 前1根ZT + 当前成交量 < 前1根成交量
#                         if latest_idx < 1:
#                             return False
#                         prev1 = stock_data[latest_idx - 1]
#                         return prev1.get('ZT', False) and (latest_k['volume'] < prev1['volume'])

#                     def check_cond2():
#                         # COND2: 当前ZT + 前1根YZB + 前2根ZT + 当前成交量 < 前2根成交量
#                         if latest_idx < 2:
#                             return False
#                         prev1 = stock_data[latest_idx - 1]
#                         prev2 = stock_data[latest_idx - 2]
#                         return prev1.get('YZB', False) and prev2.get('ZT', False) and (latest_k['volume'] < prev2['volume'])

#                     def check_cond3():
#                         # COND3: 当前ZT + 前1-2根YZB + 前3根ZT + 成交量 < 前3根
#                         if latest_idx < 3:
#                             return False
#                         prev1 = stock_data[latest_idx - 1]
#                         prev2 = stock_data[latest_idx - 2]
#                         prev3 = stock_data[latest_idx - 3]
#                         return all([prev1.get('YZB'), prev2.get('YZB')]) and prev3.get('ZT') and (latest_k['volume'] < prev3['volume'])

#                     def check_cond4():
#                         # COND4: 当前ZT + 前1-3根YZB + 前4根ZT + 成交量 < 前4根
#                         if latest_idx < 4:
#                             return False
#                         prevs_yzb = [stock_data[latest_idx - i] for i in range(1, 4)]
#                         prev4 = stock_data[latest_idx - 4]
#                         return all(yzb.get('YZB') for yzb in prevs_yzb) and prev4.get('ZT') and (latest_k['volume'] < prev4['volume'])

#                     def check_cond5():
#                         # COND5: 当前ZT + 前1-4根YZB + 前5根ZT + 成交量 < 前5根
#                         if latest_idx < 5:
#                             return False
#                         prevs_yzb = [stock_data[latest_idx - i] for i in range(1, 5)]
#                         prev5 = stock_data[latest_idx - 5]
#                         return all(yzb.get('YZB') for yzb in prevs_yzb) and prev5.get('ZT') and (latest_k['volume'] < prev5['volume'])

#                     def check_cond6():
#                         # COND6: 当前ZT + 前1-5根YZB + 前6根ZT + 成交量 < 前6根
#                         if latest_idx < 6:
#                             return False
#                         prevs_yzb = [stock_data[latest_idx - i] for i in range(1, 6)]
#                         prev6 = stock_data[latest_idx - 6]
#                         return all(yzb.get('YZB') for yzb in prevs_yzb) and prev6.get('ZT') and (latest_k['volume'] < prev6['volume'])

#                     def check_cond7():
#                         # COND7: 当前ZT + 前1-6根YZB + 前7根ZT + 成交量 < 前7根
#                         if latest_idx < 7:
#                             return False
#                         prevs_yzb = [stock_data[latest_idx - i] for i in range(1, 7)]
#                         prev7 = stock_data[latest_idx - 7]
#                         return all(yzb.get('YZB') for yzb in prevs_yzb) and prev7.get('ZT') and (latest_k['volume'] < prev7['volume'])

#                     def check_cond8():
#                         # COND8: 当前ZT + 前1-7根YZB + 前8根ZT + 成交量 < 前8根
#                         if latest_idx < 8:
#                             return False
#                         prevs_yzb = [stock_data[latest_idx - i] for i in range(1, 8)]
#                         prev8 = stock_data[latest_idx - 8]
#                         return all(yzb.get('YZB') for yzb in prevs_yzb) and prev8.get('ZT') and (latest_k['volume'] < prev8['volume'])

#                     def check_cond9():
#                         # COND9: 当前ZT + 前1-8根YZB + 前9根ZT + 成交量 < 前9根
#                         if latest_idx < 9:
#                             return False
#                         prevs_yzb = [stock_data[latest_idx - i] for i in range(1, 9)]
#                         prev9 = stock_data[latest_idx - 9]
#                         return all(yzb.get('YZB') for yzb in prevs_yzb) and prev9.get('ZT') and (latest_k['volume'] < prev9['volume'])

#                     # 检查是否满足任一条件
#                     if any([check_cond1(), check_cond2(), check_cond3(), check_cond4(),
#                             check_cond5(), check_cond6(), check_cond7(), check_cond8(), check_cond9()]):
#                         logging.info(f"{full_code} 满足选股条件！")
#                         # 仅首次触发时写入文件
#                         if full_code not in self.triggered_stocks:
#                             self.write_to_blk_files(market, stock_code)
#                             self.triggered_stocks.add(full_code)
#                 else:
#                     logging.warning(f"获取 {full_code} 数据失败")
                    
#             except Exception as e:
#                 logging.error(f"处理 {full_code} 时发生错误: {e}")
#         return 


    
#     def get_stock_data_from_redis(self, full_code):
#         """
#         从Redis获取某只股票的所有数据
        
#         Args:
#             full_code: 完整股票代码
            
#         Returns:
#             list: 股票数据列表
#         """
#         try:
#             redis_key = f"shishi:{full_code}"
#             data_list = self.redis_client.lrange(redis_key, 0, -1)
            
#             result = []
#             for data_json in data_list:
#                 data_item = json.loads(data_json)
#                 result.append(data_item)
            
#             return result
#         except Exception as e:
#             logging.error(f"从Redis获取 {full_code} 数据失败: {e}")
#             return []
    
#     def run(self, interval_seconds=2):
#         """
#         运行数据收集器
        
#         Args:
#             interval_seconds: 更新间隔（秒）
#         """
        
#         logging.info(f"开始定时数据收集，间隔: {interval_seconds}秒")
        
#         # 立即执行一次
#         self.update_all_stocks()
        
#         # 设置定时任务
#         # schedule.every(interval_seconds).seconds.do(self.update_all_stocks)
        
#         # try:
#         #     while True:
#         #         schedule.run_pending()
#         #         time.sleep(1)
#         # except KeyboardInterrupt:
#         #     logging.info("程序被用户中断")
#         # except Exception as e:
#         #     logging.error(f"程序运行出错: {e}")


# def main():
#     """
#     主函数
#     """
#     # 配置参数
#     BLOB_FILE_PATH = r"D:\zd_hbzq\T0002\blocknew\ZB.blk"
#     UPDATE_INTERVAL = 1       # 更新间隔（秒）
    
#     # 检查blk文件是否存在
#     if not os.path.exists(BLOB_FILE_PATH):
#         logging.error(f"blk文件不存在: {BLOB_FILE_PATH}")
#         return
    
#     # 创建并运行数据收集器
#     collector = StockDataCollector(
#         blk_file_path=BLOB_FILE_PATH,
#     )
    
#     # 运行收集器
#     collector.run(interval_seconds=UPDATE_INTERVAL)


# if __name__ == "__main__":
#     main()