import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import matplotlib.colors as mcolors

# 1. 修正路径和编码（关键：用 r 前缀 + 中文编码 gbk）
file_path = r'D:\feihu\gu_piao_to_redis\20251101历史成交查询.xlsx'

# 读取 .xls 文件（用 xlrd 引擎，需先安装：pip install xlrd==1.2.0）
df = pd.read_excel(
    file_path,
    engine='xlrd',  # 适配 .xls 格式
    encoding='gbk'  # 中文编码，解决解码错误
)

# 提取E列（证券名称）的唯一值（确保表头正确，若表头不是“证券名称”需修改）
unique_names = df['证券名称'].drop_duplicates().tolist()
if '证券名称' in unique_names:
    unique_names.remove('证券名称')

# 2. 分配颜色
color_list = list(mcolors.TABLEAU_COLORS.values())  # 区分度高的配色
name_color_map = {}
for i, name in enumerate(unique_names):
    color_code = color_list[i % len(color_list)].replace('#', '')  # 转换为RGB格式
    name_color_map[name] = color_code

# 3. 用 openpyxl 处理格式（注意：.xls 格式需先另存为 .xlsx 才能用 openpyxl 编辑，这里直接处理读取的文件）
# 若文件是 .xls 格式，建议先手动另存为 .xlsx（Excel 中“另存为”选择 .xlsx），再修改路径为 .xlsx
# 以下代码假设文件已转为 .xlsx 格式（若仍是 .xls，需先转换）
wb = load_workbook(file_path.replace('.xls', '.xlsx'))  # 读取 .xlsx 文件
ws = wb.active

# 遍历E列（从第2行开始，跳过表头）
for row in range(2, ws.max_row + 1):
    cell_value = ws[f'E{row}'].value
    if cell_value in name_color_map:
        fill = PatternFill(
            start_color=name_color_map[cell_value],
            end_color=name_color_map[cell_value],
            fill_type='solid'
        )
        ws[f'E{row}'].fill = fill

# 4. 保存结果
output_path = r'D:\feihu\gu_piao_to_redis\20251101历史成交查询_带颜色标记.xlsx'
wb.save(output_path)

print(f"处理完成，新文件保存至：{output_path}")